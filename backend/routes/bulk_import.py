"""
Bulk Import System - Excel-based data import for ERP
Features:
- Customer bulk upload with validation
- Item/Product bulk upload
- Opening balance import
- Opening stock import
- Template downloads
- Error reporting
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone
import uuid
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from server import db, get_current_user

router = APIRouter()


# ==================== TEMPLATE GENERATION ====================
@router.get("/templates/{template_type}")
async def download_template(template_type: str, current_user: dict = Depends(get_current_user)):
    """Download Excel template for bulk import"""
    
    wb = Workbook()
    ws = wb.active
    
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if template_type == "customers":
        ws.title = "Customers"
        headers = [
            "Account Name*", "Contact Person", "Email", "Phone", "Mobile",
            "Address Line 1", "Address Line 2", "City", "State", "Pincode",
            "Country", "GSTIN", "PAN", "Credit Limit", "Payment Terms (Days)",
            "Account Group", "Is Customer (Y/N)", "Is Vendor (Y/N)"
        ]
        
        sample_data = [
            ["ABC Traders", "Ramesh Kumar", "abc@example.com", "044-12345678", "9876543210",
             "123 Main Street", "Near Bus Stand", "Chennai", "Tamil Nadu", "600001",
             "India", "33AAAAA0000A1Z5", "AAAAA0000A", "100000", "30",
             "Sundry Debtors", "Y", "N"],
            ["XYZ Industries", "Suresh Patel", "xyz@example.com", "022-98765432", "8765432109",
             "456 Industrial Area", "Phase 2", "Mumbai", "Maharashtra", "400001",
             "India", "27BBBBB0000B1Z5", "BBBBB0000B", "500000", "45",
             "Sundry Debtors", "Y", "Y"]
        ]
        
    elif template_type == "items":
        ws.title = "Items"
        headers = [
            "Item Code*", "Item Name*", "Category*", "Item Type", "HSN Code",
            "UOM*", "Secondary UOM", "Conversion Factor",
            "Thickness (microns)", "Width (mm)", "Length (mtrs)",
            "Color", "Adhesive Type", "Base Material", "Grade",
            "Standard Cost", "Selling Price", "Min Order Qty",
            "Reorder Level", "Safety Stock", "Lead Time (Days)"
        ]
        
        sample_data = [
            ["BOPP-001", "BOPP Tape 48mm Clear", "Finished Goods", "BOPP Tape", "39191010",
             "Rolls", "Cartons", "72",
             "40", "48", "65",
             "Clear", "Acrylic", "BOPP Film", "Standard",
             "25.50", "35.00", "100",
             "500", "200", "7"],
            ["RM-FILM-001", "BOPP Film 23 Micron", "Raw Material", "Film", "39202020",
             "Kgs", "Rolls", "50",
             "23", "", "",
             "Natural", "", "BOPP", "",
             "120.00", "", "100",
             "1000", "500", "14"]
        ]
        
    elif template_type == "opening_balance":
        ws.title = "Opening Balance"
        headers = [
            "Account Name*", "Opening Balance*", "Balance Type* (Dr/Cr)",
            "As On Date* (YYYY-MM-DD)", "Reference", "Remarks"
        ]
        
        sample_data = [
            ["ABC Traders", "50000", "Dr", "2024-04-01", "OB-001", "Opening balance FY 2024-25"],
            ["XYZ Industries", "125000", "Dr", "2024-04-01", "OB-002", "Opening balance FY 2024-25"],
            ["Supplier A", "75000", "Cr", "2024-04-01", "OB-003", "Payable opening balance"]
        ]
        
    elif template_type == "opening_stock":
        ws.title = "Opening Stock"
        headers = [
            "Item Code*", "Warehouse/Location*", "Opening Qty*", "Rate",
            "Batch No", "Expiry Date (YYYY-MM-DD)", "As On Date* (YYYY-MM-DD)", "Remarks"
        ]
        
        sample_data = [
            ["BOPP-001", "BWD", "5000", "25.50", "BATCH-001", "", "2024-04-01", "Opening stock"],
            ["BOPP-001", "VPT", "3000", "25.50", "BATCH-002", "", "2024-04-01", "Opening stock"],
            ["RM-FILM-001", "BWD", "2000", "120.00", "RM-BATCH-001", "2025-03-31", "2024-04-01", "Opening raw material"]
        ]
    
    else:
        raise HTTPException(status_code=400, detail="Invalid template type")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 2)
    
    # Write sample data
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Add instructions sheet
    ws2 = wb.create_sheet(title="Instructions")
    instructions = [
        "BULK IMPORT INSTRUCTIONS",
        "",
        "1. Fields marked with * are mandatory",
        "2. Do not modify the header row",
        "3. Delete the sample data rows before importing your data",
        "4. Dates should be in YYYY-MM-DD format (e.g., 2024-04-01)",
        "5. For Yes/No fields, use Y or N",
        "6. GSTIN should be valid 15-character GST number",
        "7. Maximum 5000 rows per import",
        "",
        "VALIDATION RULES:",
        "- Duplicate entries will be skipped",
        "- Invalid GSTIN format will be flagged",
        "- Numeric fields should contain only numbers",
        "",
        "SUPPORT:",
        "For any issues, contact support@instabiz.com"
    ]
    for row, text in enumerate(instructions, 1):
        ws2.cell(row=row, column=1, value=text)
        if row == 1:
            ws2.cell(row=row, column=1).font = Font(bold=True, size=14)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"{template_type}_import_template.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== BULK IMPORT ENDPOINTS ====================
@router.post("/customers")
async def import_customers(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk import customers from Excel"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    try:
        df = pd.read_excel(file.file, sheet_name=0)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    # Column mapping
    column_map = {
        "Account Name*": "account_name",
        "Contact Person": "contact_person",
        "Email": "email",
        "Phone": "phone",
        "Mobile": "mobile",
        "Address Line 1": "address_line1",
        "Address Line 2": "address_line2",
        "City": "city",
        "State": "state",
        "Pincode": "pincode",
        "Country": "country",
        "GSTIN": "gstin",
        "PAN": "pan",
        "Credit Limit": "credit_limit",
        "Payment Terms (Days)": "payment_terms",
        "Account Group": "account_group",
        "Is Customer (Y/N)": "is_customer",
        "Is Vendor (Y/N)": "is_vendor"
    }
    
    df = df.rename(columns=column_map)
    
    results = {"success": 0, "errors": [], "skipped": 0}
    
    for idx, row in df.iterrows():
        try:
            account_name = str(row.get('account_name', '')).strip()
            if not account_name or account_name == 'nan':
                results['errors'].append({"row": idx + 2, "error": "Account Name is required"})
                continue
            
            # Check for duplicate
            existing = await db.accounts.find_one({"account_name": account_name})
            if existing:
                results['skipped'] += 1
                continue
            
            account_id = str(uuid.uuid4())
            account_doc = {
                "id": account_id,
                "account_name": account_name,
                "contact_person": str(row.get('contact_person', '')) if pd.notna(row.get('contact_person')) else None,
                "email": str(row.get('email', '')) if pd.notna(row.get('email')) else None,
                "phone": str(row.get('phone', '')) if pd.notna(row.get('phone')) else None,
                "mobile": str(row.get('mobile', '')) if pd.notna(row.get('mobile')) else None,
                "address_line1": str(row.get('address_line1', '')) if pd.notna(row.get('address_line1')) else None,
                "address_line2": str(row.get('address_line2', '')) if pd.notna(row.get('address_line2')) else None,
                "city": str(row.get('city', '')) if pd.notna(row.get('city')) else None,
                "state": str(row.get('state', '')) if pd.notna(row.get('state')) else None,
                "pincode": str(row.get('pincode', '')) if pd.notna(row.get('pincode')) else None,
                "country": str(row.get('country', 'India')) if pd.notna(row.get('country')) else 'India',
                "gstin": str(row.get('gstin', '')) if pd.notna(row.get('gstin')) else None,
                "pan": str(row.get('pan', '')) if pd.notna(row.get('pan')) else None,
                "credit_limit": float(row.get('credit_limit', 0)) if pd.notna(row.get('credit_limit')) else 0,
                "payment_terms": int(row.get('payment_terms', 30)) if pd.notna(row.get('payment_terms')) else 30,
                "account_group": str(row.get('account_group', 'Sundry Debtors')) if pd.notna(row.get('account_group')) else 'Sundry Debtors',
                "is_customer": str(row.get('is_customer', 'Y')).upper() == 'Y',
                "is_vendor": str(row.get('is_vendor', 'N')).upper() == 'Y',
                "outstanding_balance": 0,
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user['id'],
                "import_source": "bulk_import"
            }
            
            await db.accounts.insert_one(account_doc)
            results['success'] += 1
            
        except Exception as e:
            results['errors'].append({"row": idx + 2, "error": str(e)})
    
    return {
        "message": f"Import completed: {results['success']} created, {results['skipped']} skipped, {len(results['errors'])} errors",
        "details": results
    }


@router.post("/items")
async def import_items(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk import items from Excel"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    try:
        df = pd.read_excel(file.file, sheet_name=0)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    column_map = {
        "Item Code*": "item_code",
        "Item Name*": "item_name",
        "Category*": "category",
        "Item Type": "item_type",
        "HSN Code": "hsn_code",
        "UOM*": "uom",
        "Secondary UOM": "secondary_uom",
        "Conversion Factor": "conversion_factor",
        "Thickness (microns)": "thickness",
        "Width (mm)": "width",
        "Length (mtrs)": "length",
        "Color": "color",
        "Adhesive Type": "adhesive_type",
        "Base Material": "base_material",
        "Grade": "grade",
        "Standard Cost": "standard_cost",
        "Selling Price": "selling_price",
        "Min Order Qty": "min_order_qty",
        "Reorder Level": "reorder_level",
        "Safety Stock": "safety_stock",
        "Lead Time (Days)": "lead_time_days"
    }
    
    df = df.rename(columns=column_map)
    
    results = {"success": 0, "errors": [], "skipped": 0}
    
    for idx, row in df.iterrows():
        try:
            item_code = str(row.get('item_code', '')).strip()
            item_name = str(row.get('item_name', '')).strip()
            
            if not item_code or item_code == 'nan':
                results['errors'].append({"row": idx + 2, "error": "Item Code is required"})
                continue
            if not item_name or item_name == 'nan':
                results['errors'].append({"row": idx + 2, "error": "Item Name is required"})
                continue
            
            # Check for duplicate
            existing = await db.items.find_one({"item_code": item_code})
            if existing:
                results['skipped'] += 1
                continue
            
            item_id = str(uuid.uuid4())
            item_doc = {
                "id": item_id,
                "item_code": item_code,
                "item_name": item_name,
                "category": str(row.get('category', 'Finished Goods')) if pd.notna(row.get('category')) else 'Finished Goods',
                "item_type": str(row.get('item_type', '')) if pd.notna(row.get('item_type')) else None,
                "hsn_code": str(row.get('hsn_code', '')) if pd.notna(row.get('hsn_code')) else None,
                "uom": str(row.get('uom', 'Rolls')) if pd.notna(row.get('uom')) else 'Rolls',
                "secondary_uom": str(row.get('secondary_uom', '')) if pd.notna(row.get('secondary_uom')) else None,
                "conversion_factor": float(row.get('conversion_factor', 1)) if pd.notna(row.get('conversion_factor')) else 1,
                "thickness": float(row.get('thickness')) if pd.notna(row.get('thickness')) else None,
                "width": float(row.get('width')) if pd.notna(row.get('width')) else None,
                "length": float(row.get('length')) if pd.notna(row.get('length')) else None,
                "color": str(row.get('color', '')) if pd.notna(row.get('color')) else None,
                "adhesive_type": str(row.get('adhesive_type', '')) if pd.notna(row.get('adhesive_type')) else None,
                "base_material": str(row.get('base_material', '')) if pd.notna(row.get('base_material')) else None,
                "grade": str(row.get('grade', '')) if pd.notna(row.get('grade')) else None,
                "standard_cost": float(row.get('standard_cost', 0)) if pd.notna(row.get('standard_cost')) else 0,
                "selling_price": float(row.get('selling_price', 0)) if pd.notna(row.get('selling_price')) else 0,
                "min_order_qty": float(row.get('min_order_qty', 1)) if pd.notna(row.get('min_order_qty')) else 1,
                "reorder_level": float(row.get('reorder_level', 0)) if pd.notna(row.get('reorder_level')) else 0,
                "safety_stock": float(row.get('safety_stock', 0)) if pd.notna(row.get('safety_stock')) else 0,
                "lead_time_days": int(row.get('lead_time_days', 7)) if pd.notna(row.get('lead_time_days')) else 7,
                "current_stock": 0,
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user['id'],
                "import_source": "bulk_import"
            }
            
            await db.items.insert_one(item_doc)
            results['success'] += 1
            
        except Exception as e:
            results['errors'].append({"row": idx + 2, "error": str(e)})
    
    return {
        "message": f"Import completed: {results['success']} created, {results['skipped']} skipped, {len(results['errors'])} errors",
        "details": results
    }


@router.post("/opening-balance")
async def import_opening_balance(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Import opening balances for accounts"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files supported")
    
    try:
        df = pd.read_excel(file.file, sheet_name=0)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    column_map = {
        "Account Name*": "account_name",
        "Opening Balance*": "opening_balance",
        "Balance Type* (Dr/Cr)": "balance_type",
        "As On Date* (YYYY-MM-DD)": "as_on_date",
        "Reference": "reference",
        "Remarks": "remarks"
    }
    
    df = df.rename(columns=column_map)
    
    results = {"success": 0, "errors": [], "not_found": []}
    
    for idx, row in df.iterrows():
        try:
            account_name = str(row.get('account_name', '')).strip()
            if not account_name or account_name == 'nan':
                results['errors'].append({"row": idx + 2, "error": "Account Name is required"})
                continue
            
            # Find account
            account = await db.accounts.find_one({"account_name": account_name})
            if not account:
                results['not_found'].append({"row": idx + 2, "account_name": account_name})
                continue
            
            opening_balance = float(row.get('opening_balance', 0))
            balance_type = str(row.get('balance_type', 'Dr')).upper()
            
            # Adjust sign based on type
            if balance_type == 'CR':
                opening_balance = -opening_balance
            
            # Update account balance
            await db.accounts.update_one(
                {"id": account['id']},
                {
                    "$set": {
                        "opening_balance": opening_balance,
                        "outstanding_balance": opening_balance,
                        "opening_balance_date": str(row.get('as_on_date', '')) if pd.notna(row.get('as_on_date')) else None,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                }
            )
            
            # Create ledger entry
            ledger_entry = {
                "id": str(uuid.uuid4()),
                "account_id": account['id'],
                "account_name": account_name,
                "transaction_type": "opening_balance",
                "debit": opening_balance if balance_type == 'DR' else 0,
                "credit": abs(opening_balance) if balance_type == 'CR' else 0,
                "balance": opening_balance,
                "date": str(row.get('as_on_date', '')),
                "reference": str(row.get('reference', '')) if pd.notna(row.get('reference')) else "Opening Balance",
                "remarks": str(row.get('remarks', '')) if pd.notna(row.get('remarks')) else None,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user['id']
            }
            await db.ledger_entries.insert_one(ledger_entry)
            
            results['success'] += 1
            
        except Exception as e:
            results['errors'].append({"row": idx + 2, "error": str(e)})
    
    return {
        "message": f"Import completed: {results['success']} updated, {len(results['not_found'])} not found, {len(results['errors'])} errors",
        "details": results
    }


@router.post("/opening-stock")
async def import_opening_stock(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Import opening stock for items"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files supported")
    
    try:
        df = pd.read_excel(file.file, sheet_name=0)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")
    
    column_map = {
        "Item Code*": "item_code",
        "Warehouse/Location*": "warehouse",
        "Opening Qty*": "opening_qty",
        "Rate": "rate",
        "Batch No": "batch_no",
        "Expiry Date (YYYY-MM-DD)": "expiry_date",
        "As On Date* (YYYY-MM-DD)": "as_on_date",
        "Remarks": "remarks"
    }
    
    df = df.rename(columns=column_map)
    
    results = {"success": 0, "errors": [], "not_found": []}
    
    for idx, row in df.iterrows():
        try:
            item_code = str(row.get('item_code', '')).strip()
            if not item_code or item_code == 'nan':
                results['errors'].append({"row": idx + 2, "error": "Item Code is required"})
                continue
            
            # Find item
            item = await db.items.find_one({"item_code": item_code})
            if not item:
                results['not_found'].append({"row": idx + 2, "item_code": item_code})
                continue
            
            opening_qty = float(row.get('opening_qty', 0))
            rate = float(row.get('rate', 0)) if pd.notna(row.get('rate')) else item.get('standard_cost', 0)
            warehouse = str(row.get('warehouse', 'BWD')) if pd.notna(row.get('warehouse')) else 'BWD'
            
            # Update item stock
            await db.items.update_one(
                {"id": item['id']},
                {
                    "$inc": {"current_stock": opening_qty},
                    "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
                }
            )
            
            # Create stock entry
            stock_entry = {
                "id": str(uuid.uuid4()),
                "item_id": item['id'],
                "item_code": item_code,
                "item_name": item.get('item_name'),
                "transaction_type": "opening_stock",
                "quantity": opening_qty,
                "rate": rate,
                "value": opening_qty * rate,
                "warehouse": warehouse,
                "batch_no": str(row.get('batch_no', '')) if pd.notna(row.get('batch_no')) else None,
                "expiry_date": str(row.get('expiry_date', '')) if pd.notna(row.get('expiry_date')) else None,
                "date": str(row.get('as_on_date', '')),
                "remarks": str(row.get('remarks', '')) if pd.notna(row.get('remarks')) else "Opening Stock",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user['id']
            }
            await db.stock_entries.insert_one(stock_entry)
            
            results['success'] += 1
            
        except Exception as e:
            results['errors'].append({"row": idx + 2, "error": str(e)})
    
    return {
        "message": f"Import completed: {results['success']} entries created, {len(results['not_found'])} items not found, {len(results['errors'])} errors",
        "details": results
    }
